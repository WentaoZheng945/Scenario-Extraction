import os
from tqdm import tqdm
import openpyxl
import numpy as np
import pandas as pd
import xml.etree.ElementTree as ET


def tag_nds(nds_path):
    '''
    该方法用于自动化实现NDS场景有关场景危险度MTTC及场景对手车数量的标签值计算

    Input：NDS场景XML文件路径
    Output：场景对手车数量、场景危险度（以最小MTTC衡量）、主车行为、对手车行为
    '''
    nds_data = open(nds_path).read()
    nds_root = ET.fromstring(nds_data)
    sample = 1 / 10  # 上海NDS场景的轨迹采样率
    speed_list = []
    for speed in nds_root.iter('Speed'):  # 获得场景所有参与者初始速度
        speed_list.append(float(speed.attrib['Value']))
    x, y, x_ego, y_ego = [[] for i in range(4)]
    for path in nds_root.iter('PathShape'):  # 获得场景所有参与者对应的轨迹信息
        x_list, y_list = [[] for i in range(2)]
        for waypoint in path:
            x_list.append(float(waypoint.attrib['X']))
            y_list.append(float(waypoint.attrib['Y']))
        if x_ego:
            x.append(x_list)  # 二维列表，用于储存每辆车对应的xy信息
            y.append(y_list)
        else:
            x_ego = x_list
            y_ego = y_list

    # TODO: 遍历主车轨迹，识别前车，并计算mttc
    # 前车的定义为与本车横向距离小于2m的前方空间内，纵向距离最近的同向行驶车辆
    mttc_list = []
    id_list = []
    frame_list = []
    for i, ego_x in enumerate(x_ego):
        if i != 0 and i != len(x_ego) - 1:  # 除去首尾两个轨迹点（无法计算速度、加速度）
            ego_y = y_ego[i]
            temp_x = np.array([temp[i] for temp in x])  # 取出对应时刻所有车对应的位置
            temp_y = np.array([temp[i] for temp in y])
            id1 = np.where(temp_x > ego_x)[0]  # 位于主车前方
            id2 = np.where((ego_y - 1 < temp_y) & (temp_y < ego_y + 1))[0]  # 主车前方横向2m内
            car_id = list(set(id1).intersection(set(id2)))  # 取交集，获得该帧下的前车索引
            # 分别计算对应的MTTC
            # 当前位置速度为下一帧相邻两点平均速度，加速度为上一帧及下一帧三个点对应两段平均速度的平均加速度
            for id in car_id:
                car_x = x[id]
                car_y = y[id]
                dist = ((ego_x - car_x[i])**2 + (ego_y - car_y[i])**2)**0.5
                ego_v = (((ego_x - x_ego[i + 1])**2 + (ego_y - y_ego[i + 1])**2)**0.5) / sample
                car_v = (((car_x[i] - car_x[i + 1])**2 + (car_y[i] - car_y[i + 1])**2)**0.5) / sample
                delta_v = ego_v - car_v
                delta_v2 = (ego_v - car_v)**2
                ego_a = (ego_v - (((ego_x - x_ego[i - 1])**2 + (ego_y - y_ego[i - 1])**2)**0.5) / sample) / sample
                car_a = (car_v - (((car_x[i - 1] - car_x[i])**2 + (car_y[i - 1] - car_y[i])**2)**0.5) / sample) / sample
                delta_a = ego_a - car_a
                if delta_a > 0:  # 与原文不同（原文！=0即可），在此思路下，根号内容易计算出负数（dist过大），从而得到虚数
                    t1 = (delta_v * (-1) - (delta_v2 + 2 * delta_a * dist)**0.5) / delta_a
                    t2 = (delta_v * (-1) + (delta_v2 + 2 * delta_a * dist)**0.5) / delta_a
                    if t1 > 0 and t2 > 0:
                        mttc = min(t1, t2)
                    elif t1 * t2 <= 0 and max(t1, t2) > 0:
                        mttc = max(t1, t2)
                elif delta_a == 0 and delta_v > 0:
                    mttc = dist / delta_v
                else:  # 速度差小于零并且不存在加速度差的情况，不存在碰撞风险
                    mttc = float('inf')
                mttc_list.append(round(mttc, 2))
                id_list.append(id)
                frame_list.append(i)
    num_challengingcar = len(x)
    if mttc_list:
        min_ttc = min(mttc_list)
        interact_id = id_list[mttc_list.index(min_ttc)]  # 与ego交互的对手车在xy二维列表中对应的索引
        highrisk_frame = frame_list[mttc_list.index(min_ttc)]  # min_MTTC对应的高风险时刻（列表索引）

        # TODO：分析主车与对手车轨迹，识别其交互动作
        # 交互动作识别（初始与高风险时刻的比较）
        # 以场景类型做区分，对于紧密跟驰场景，关注其速度变化（前后车加减速情况）；换道场景，关注cutin情况
        x_interactor = x[interact_id]
        y_interactor = y[interact_id]
        a_interactor = []
        if 'follow' in nds_path:  # 紧密跟驰场景，主要关注前车突然减速所导致的危险情况，减速度阈值选择2m/s^2
            for i, interactor_x in enumerate(x_interactor):
                if i != 0 and i != len(x_interactor) - 1:
                    car_v = (((x_interactor[i] - x_interactor[i + 1])**2 + (y_interactor[i] - y_interactor[i + 1])**2)**0.5) / sample
                    car_a = (car_v - (((x_interactor[i - 1] - x_interactor[i])**2 + (y_interactor[i - 1] - y_interactor[i])**2)**0.5) / sample) / sample
                    a_interactor.append(car_a)
            if max(a_interactor) >= 2:
                behavior_car = '急减速'
                behavior_ego = '直行'
            else:
                behavior_car = '直行'
                behavior_ego = '直行'
        else:  # 存在换道行为的场景，以1.2m范围的横向空间为阈值对换道/车道内摇摆进行区分
            if y_interactor[highrisk_frame] - y_interactor[0] > 0.6:
                behavior_car = '右侧cutin'
            elif y_interactor[highrisk_frame] - y_interactor[0] < -0.6:
                behavior_car = '左侧cutin'
            else:
                behavior_car = '直行'
            if y_ego[highrisk_frame] - y_ego[0] > 0.6:
                behavior_ego = '向左换道'
            elif y_ego[highrisk_frame] - y_interactor[0] < -0.6:
                behavior_ego = '向右换道'
            else:
                behavior_ego = '直行'
    else:
        min_ttc = ''
        behavior_car = ''
        behavior_ego = ''
    return num_challengingcar, min_ttc, behavior_ego, behavior_car


def tag_highD(csv_path):
    '''
    该方法用于自动化实现highD场景有关标签值计算

    Input：NDS场景XML文件路径
    Output：主车行为、对手车行为
    '''
    sample = 1 / 25
    df = pd.read_csv(csv_path)
    grouped = df.groupby(['id'], sort=False)  # 将df按照车辆id进行group
    count = 0
    x, y = [[] for i in range(2)]
    for group_id, rows in grouped:
        if count == 0:  # ego
            x_ego = rows['x'].values.tolist()
            y_ego = rows['y'].values.tolist()
            count += 1
        else:
            x.append(rows['x'].values.tolist())
            y.append(rows['y'].values.tolist())

    # TODO: 遍历主车轨迹，识别前车，并计算mttc
    # 前车的定义为与本车横向距离小于2m的前方空间内，纵向距离最近的同向行驶车辆
    mttc_list = []
    id_list = []
    frame_list = []
    for i, ego_x in enumerate(x_ego):
        if i != 0 and i != len(x_ego) - 1:  # 除去首尾两个轨迹点（无法计算速度、加速度）
            ego_y = y_ego[i]
            temp_x = np.array([temp[i] for temp in x])  # 取出对应时刻所有车对应的位置
            temp_y = np.array([temp[i] for temp in y])
            if x_ego[0] < x_ego[-1]:  # 下行方向
                id1 = np.where(temp_x > ego_x)[0]  # 位于主车前方
                id2 = np.where((ego_y - 1 < temp_y) & (temp_y < ego_y + 1))[0]  # 主车前方横向2m内
            else:
                id1 = np.where(temp_x < ego_x)[0]  # 位于主车前方
                id2 = np.where((ego_y - 1 < temp_y) & (temp_y < ego_y + 1))[0]  # 主车前方横向2m内
            car_id = list(set(id1).intersection(set(id2)))  # 取交集，获得该帧下的前车索引
            # 分别计算对应的MTTC
            # 当前位置速度为下一帧相邻两点平均速度，加速度为上一帧及下一帧三个点对应两段平均速度的平均加速度
            for id in car_id:
                car_x = x[id]
                car_y = y[id]
                dist = ((ego_x - car_x[i])**2 + (ego_y - car_y[i])**2)**0.5
                ego_v = (((ego_x - x_ego[i + 1])**2 + (ego_y - y_ego[i + 1])**2)**0.5) / sample
                car_v = (((car_x[i] - car_x[i + 1])**2 + (car_y[i] - car_y[i + 1])**2)**0.5) / sample
                delta_v = ego_v - car_v
                delta_v2 = (ego_v - car_v)**2
                ego_a = (ego_v - (((ego_x - x_ego[i - 1])**2 + (ego_y - y_ego[i - 1])**2)**0.5) / sample) / sample
                car_a = (car_v - (((car_x[i - 1] - car_x[i])**2 + (car_y[i - 1] - car_y[i])**2)**0.5) / sample) / sample
                delta_a = ego_a - car_a
                if delta_a > 0:  # 与原文不同（原文！=0即可），在此思路下，根号内容易计算出负数（dist过大），从而得到虚数
                    t1 = (delta_v * (-1) - (delta_v2 + 2 * delta_a * dist)**0.5) / delta_a
                    t2 = (delta_v * (-1) + (delta_v2 + 2 * delta_a * dist)**0.5) / delta_a
                    if t1 > 0 and t2 > 0:
                        mttc = min(t1, t2)
                    elif t1 * t2 <= 0 and max(t1, t2) > 0:
                        mttc = max(t1, t2)
                elif delta_a == 0 and delta_v > 0:
                    mttc = dist / delta_v
                else:  # 速度差小于零并且不存在加速度差的情况，不存在碰撞风险
                    mttc = float('inf')
                mttc_list.append(round(mttc, 2))
                id_list.append(id)
                frame_list.append(i)
    num_challengingcar = len(x)
    if mttc_list:
        min_ttc = min(mttc_list)
        interact_id = id_list[mttc_list.index(min_ttc)]  # 与ego交互的对手车在xy二维列表中对应的索引
        highrisk_frame = frame_list[mttc_list.index(min_ttc)]  # min_MTTC对应的高风险时刻（列表索引）

        # TODO：分析主车与对手车轨迹，识别其交互动作
        # 交互动作识别（初始与高风险时刻的比较）
        # 以场景类型做区分，对于紧密跟驰场景，关注其速度变化（前后车加减速情况）；换道场景，关注cutin情况
        x_interactor = x[interact_id]
        y_interactor = y[interact_id]
        a_interactor = []
        if 'follow' in csv_path:  # 紧密跟驰场景，主要关注前车突然减速所导致的危险情况，减速度阈值选择2m/s^2
            for i, interactor_x in enumerate(x_interactor):
                if i != 0 and i != len(x_interactor) - 1:
                    car_v = (((x_interactor[i] - x_interactor[i + 1])**2 + (y_interactor[i] - y_interactor[i + 1])**2)**0.5) / sample
                    car_a = (car_v - (((x_interactor[i - 1] - x_interactor[i])**2 + (y_interactor[i - 1] - y_interactor[i])**2)**0.5) / sample) / sample
                    a_interactor.append(car_a)
            if max(a_interactor) >= 2:
                behavior_car = '急减速'
                behavior_ego = '直行'
            else:
                behavior_car = '直行'
                behavior_ego = '直行'
        else:  # 存在换道行为的场景，以1.2m范围的横向空间为阈值对换道/车道内摇摆进行区分
            if y_interactor[highrisk_frame] - y_interactor[0] > 0.6:
                behavior_car = '右侧cutin'
            elif y_interactor[highrisk_frame] - y_interactor[0] < -0.6:
                behavior_car = '左侧cutin'
            else:
                behavior_car = '直行'
            if y_ego[highrisk_frame] - y_ego[0] > 0.6:
                behavior_ego = '向左换道'
            elif y_ego[highrisk_frame] - y_interactor[0] < -0.6:
                behavior_ego = '向右换道'
            else:
                behavior_ego = '直行'
    else:
        min_ttc = ''
        behavior_car = ''
        behavior_ego = ''
    return num_challengingcar, min_ttc, behavior_ego, behavior_car


if __name__ == '__main__':
    workbook = openpyxl.load_workbook('scenario_index.xlsx')
    worksheet = workbook.worksheets[0]
    names = []
    for cell in list(worksheet.columns)[0]:
        names.append(cell.value)
    # root_path = 'C:\\Users\\15251\\Desktop\\OnSite\\OpenS_for_EKT_20220114\\NDS'
    root_path = 'C:\\Users\\15251\\Desktop\\OnSite\\OpenSCENARIO\\200_for_EKT'
    for root, dirs, files in os.walk(root_path):
        for dir in tqdm(dirs):
            # nds_path = os.path.join(root, dir) + '/' + dir + '.xml'
            # num, mttc, behavior_ego, behavior_car = tag_nds(nds_path)
            csv_path = os.path.join(root, dir) + '/' + dir + '_test' + '.csv'
            num, mttc, behavior_ego, behavior_car = tag_highD(csv_path)
            if dir in names:
                index = names.index(dir)
                worksheet.cell(index + 1, 5, num)
                worksheet.cell(index + 1, 6, mttc)
                worksheet.cell(index + 1, 7, behavior_ego)
                worksheet.cell(index + 1, 8, behavior_car)
        workbook.save('scenario_index.xlsx')
        break 
